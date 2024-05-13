from openpyxl import load_workbook
import os
import pdb

def get_word_list_for_synthesis(filepath):

    print('Loading file',  filepath)
    wb = load_workbook(filepath)
    words_list = []

    for sheetname in wb.sheetnames:
        print('Loading {} sheet'.format(sheetname))    
        sheet = wb[sheetname]
        word_column_index = None
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=col).value == 'word' or sheet.cell(row=1, column=col).value == 'SyllableLow':
                word_column_index = col
                break
        if word_column_index is not None:
            for row in range(2, sheet.max_row + 1):
                word_value = sheet.cell(row=row, column=word_column_index).value
                if word_value not in words_list:
                        words_list.append(word_value)
    words_list = [word for word in words_list if word is not None]
    return words_list

def get_all_files_with_paths(folder_path):
    files_with_paths = []

    for root, dirs, files in os.walk(folder_path):
        for file in files:
            file_path = os.path.join(root, file)
            files_with_paths.append(file_path)

    return files_with_paths

def get_all_filenames_from_directory(directory_path):
    filenames = os.listdir(directory_path)
    filenames_no_extension = [os.path.splitext(filename)[0] for filename in filenames]

    return filenames_no_extension


def get_directories_in_current_folder(folder_path):    
    directories = [d for d in os.listdir(folder_path) if os.path.isdir(os.path.join(folder_path, d))]
    
    return directories


def get_word_list_with_frequency_info(filepath):
    import pandas as pd
    print('Loading file',  filepath)
    excel_data = pd.read_excel(filepath, sheet_name=None)
    frequency_group ={}
    
    for sheet_name, sheet_data in excel_data.items():
        print(f"Sheet Name: {sheet_name}")
        
        # Assuming 'word' and 'FrequencyGroup' are column names
        try:

            word_frequency_pairs = sheet_data[['word', 'FreqencyGroup']]
            
            # Iterate through rows to print each word and its frequency group
            for index, row in word_frequency_pairs.iterrows():
                word = row['word']
                if word in frequency_group:
                    continue
                frequency_group[word] = row['FreqencyGroup']

        except:
            pass 

    return frequency_group